"""Fonction utilitaires."""

import os
import logging
from collections import defaultdict
from itertools import combinations
from functools import partial
import re
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sti_loader import STILoader

KEY_COLS = ["Reference", "Requirement"]

# Configuration de base du logger
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        # logging.StreamHandler(),  # Console
        logging.FileHandler("analyse_structure.log"),  # Optionnel : fichier
    ],
)


def get_matrix_pairs(loader):
    """Regroupe les matrices par suffixe (la partie après le 1er '_')."""
    suffix_map = defaultdict(list)
    for matrix in loader.config["matrices"]:
        name = matrix["name"]
        parts = name.split("_", 1)
        if len(parts) == 2:
            suffix_map[parts[1]].append(name)

    # Pour chaque suffixe commun à plusieurs familles, génère les paires possibles
    pairs = []
    for _, name_list in suffix_map.items():
        if len(name_list) >= 2:
            for a, b in combinations(sorted(name_list), 2):
                pairs.append((a, b))

    return pairs


def nettoyer_colonnes(df):
    """Nettoie les colonnes."""
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.replace("\n", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
    )
    return df


def separer_requis(df):
    """Separe les requis"""
    if "isRequirement" not in df.columns:
        return df.copy(), pd.DataFrame()  # Tout est pris comme requis par défaut

    valeurs_requises = ["true", "vrai", "1", "yes", "requis"]
    masque_requis = (
        df["isRequirement"].astype(str).str.strip().str.lower().isin(valeurs_requises)
    )
    return df[masque_requis].copy(), df[~masque_requis].copy()


def deduplicate_and_sort(func):
    """Decorateur utilisé."""

    def wrapper(text):
        result = func(text)
        return sorted(set(result)) if isinstance(result, list) else result

    return wrapper


@deduplicate_and_sort
def extract_documents(text):
    """
    Extrait les identifiants documentaires depuis une chaîne de texte.
    Prend en charge les préfixes comme DID, CMD, SETC, PM avec suffixes éventuels.
    """
    if not isinstance(text, str):
        return []
    # pattern = r"(?:DID|CMD|PM|SETC)[0-9]{6,}(?:[-_][A-Z0-9\.]+)*"
    pattern = r"(?:DID[0-9]{10}|CMD[0-9]{6,}|PM[0-9]{6,}|SETC[0-9]{6,})"
    return re.findall(pattern, text)


def comparer_row(row, sources):
    """Comparaison avec affichage synthétique."""
    set_1 = set(row["Docs_1"])
    set_2 = set(row["Docs_2"])

    communs = set_1 & set_2
    only_1 = set_1 - communs
    only_2 = set_2 - communs

    if not only_1 and not only_2:
        return "", "Identiques"

    bloc = []
    if only_1:
        bloc.append(f"{sources[0]} : {', '.join(sorted(only_1))}")
    if only_2:
        bloc.append(f"{sources[1]} : {', '.join(sorted(only_2))}")

    if not set_1:
        statut = f"Absent dans {sources[0]}"
    elif not set_2:
        statut = f"Absent dans {sources[1]}"
    else:
        statut = "Différents"

    return "\n".join(bloc), statut


def analyser_divergences_documentaires(df, source_cols):
    """
    Applique l’analyse des divergences documentaires sur un DataFrame issu
    de la comparaison GE vs H2.
    Retourne un DataFrame regroupé par (Reference, Champ) avec documents
    extraits et divergences.
    """
    source_1, source_2 = source_cols
    champs_cibles = ["CAF_Comments", "MOP_design", "MOP_test"]
    df_docs = df[df["Champ"].isin(champs_cibles)].copy()

    # Extraction
    df_docs["Docs_1"] = df_docs[source_1].apply(extract_documents)
    df_docs["Docs_2"] = df_docs[source_2].apply(extract_documents)

    # 🔍 Supprimer les lignes sans documents de part et d’autre
    df_docs = df_docs[
        ~((df_docs["Docs_1"].str.len() == 0) & (df_docs["Docs_2"].str.len() == 0))
    ]
    if df_docs.empty:
        logging.info("Aucune divergence documentaire détectée.")
        return pd.DataFrame()

    sources = [i.split("_")[0] for i in source_cols]

    # Définir une version partielle de la fonction avec `sources` fixé
    comparer_row_with_sources = partial(comparer_row, sources=sources)

    df_docs[["Différence", "État"]] = df_docs.apply(
        comparer_row_with_sources, axis=1, result_type="expand"
    )

    # On garde uniquement les lignes divergentes
    df_divergences = df_docs[df_docs["État"] != "Identiques"].copy()

    # Regroupement par triplet pour affichage consolidé
    regrouped = (
        df_divergences.groupby(["Champ", "État", "Différence"])["Reference"]
        .apply(list)
        .reset_index()
    )

    regrouped["nb_references"] = regrouped["Reference"].apply(len)
    regrouped = regrouped.sort_values(by="nb_references", ascending=False)

    return regrouped


def normalize(df, key_cols):
    """Normalise le dataframe."""
    return df.assign(
        **{col: df[col].astype(str).str.strip() for col in key_cols if col in df}
    )


def keyset(df, key_cols):
    """Retourne l'ensemble de clés."""
    return set(map(tuple, df[key_cols].dropna().drop_duplicates().values))


def exclusive_rows(i, sets_all):
    """Gére les doublons."""
    other = set().union(*(s for j, s in enumerate(sets_all) if j != i))
    return sets_all[i] - other


def compute_sets_summary(sets_all, labels, key_cols):
    """Calcule le résumé."""
    summary = {f"Total entries in {lbl}": len(s) for lbl, s in zip(labels, sets_all)}
    common_keys = set.intersection(*sets_all)
    summary["Common entries in all matrices"] = len(common_keys)
    common_df = pd.DataFrame(common_keys, columns=key_cols)

    exclusive = {
        lbl: pd.DataFrame(list(exclusive_rows(i, sets_all)), columns=key_cols)
        for i, lbl in enumerate(labels)
    }

    for lbl in labels:
        summary[f"Entries only in {lbl}"] = len(exclusive[lbl])

    return summary, exclusive, common_df


def compute_field_diffs(cleaned, key_cols, labels, fields_to_compare):
    """Calcule les différences."""
    df1 = cleaned[0].set_index(key_cols)
    df2 = cleaned[1].set_index(key_cols)
    rows = []
    for idx in df1.index.intersection(df2.index):
        for col in fields_to_compare:
            val1 = str(df1.at[idx, col]) if col in df1.columns else ""
            val2 = str(df2.at[idx, col]) if col in df2.columns else ""
            if val1 != val2:
                row = dict(zip(key_cols, idx))
                row.update({"Champ": col, labels[0]: val1, labels[1]: val2})
                rows.append(row)
    return pd.DataFrame(rows)


def compare_matrix_entries_multi(
    dfs: list[pd.DataFrame],
    labels: list[str],
    key_cols: Optional[list[str]] = None,
    compare_fields: Optional[bool] = False,
    fields_to_compare: Optional[list[str]] = None,
):
    """
    Compare plusieurs matrices en détectant les lignes communes + divergences champ à champ.

    Args:
        dfs (list of pd.DataFrame): Matrices à comparer.
        labels (list of str): Noms des matrices.
        key_cols (list of str): Colonnes servant de clé.
        compare_fields (bool): Si True, compare les champs (colonne par colonne).
        fields_to_compare (list[str]): Colonnes à comparer si champ à champ.

    Returns:
        summary (dict): Statistiques.
        exclusive_dfs (dict[str, pd.DataFrame]): Lignes propres à chaque source.
        common_all_df (pd.DataFrame): Lignes strictement communes sur les clés.
        diffs (pd.DataFrame | None): Divergences champ à champ (si activé).
    """
    if key_cols is None:
        key_cols = list(KEY_COLS)
    if not isinstance(key_cols, list):
        raise TypeError(f"key_cols doit être une liste, reçu : {type(key_cols)}")
    if len(dfs) != len(labels):
        raise ValueError("dfs et labels doivent avoir la même longueur.")

    cleaned = [normalize(df, key_cols) for df in dfs]
    sets_all = [keyset(df, key_cols) for df in cleaned]

    summary = {f"Total entries in {lbl}": len(s) for lbl, s in zip(labels, sets_all)}
    summary["Common entries in all matrices"] = len(set.intersection(*sets_all))
    common_df = pd.DataFrame(set.intersection(*sets_all), columns=key_cols)

    summary, exclusive, common_df = compute_sets_summary(sets_all, labels, key_cols)
    diffs = (
        compute_field_diffs(cleaned, key_cols, labels, fields_to_compare)
        if compare_fields and len(dfs) == 2
        else None
    )

    return summary, exclusive, common_df, diffs


def export_df_excel(df: pd.DataFrame, path: str):
    """Travaille l'export en excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Analyse consolidée"

    # Écriture du DataFrame dans la feuille
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=str(value))

    # Écriture de l'en-tête (ligne 1)
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Ajout du filtre automatique
    ws.auto_filter.ref = f"A1:{get_column_letter(df.shape[1])}1"

    # Ajustement dynamique des largeurs par colonne
    for col_idx, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            [len(str(cell)) if cell is not None else 0 for cell in df[col]] + [len(col)]
        )
        # Option : limite haute pour éviter les colonnes trop larges
        adjusted_width = min(max_length + 1.5, 60)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(path)


def analyser_couple_matrices(matrices_cibles, path="./data/GE_H2"):
    """Analyse chaque paire de matrices STI et génère les fichiers de sortie."""
    output_dir = os.path.join(path, "output")
    os.makedirs(output_dir, exist_ok=True)

    loader = STILoader(os.path.join(path, "sti_config.yaml"))
    cols_interessees = loader.get_fields_to_compare()

    dfs_requis, _, labels = charger_et_preparer_matrices(
        matrices_cibles, loader
    )

    if len(dfs_requis) != 2:
        raise ValueError("Comparaison impossible : matrices incomplètes.")

    summary, exclusifs, commun, divergents = compare_matrix_entries_multi(
        dfs_requis,
        labels,
        key_cols=KEY_COLS,
        compare_fields=True,
        fields_to_compare=cols_interessees,
    )

    log_summary(summary)
    exporter_resultats(output_dir, exclusifs, commun, divergents, labels)
    return analyser_si_divergences(divergents, output_dir, labels)


def charger_et_preparer_matrices(matrices_cibles, loader):
    """Charge et prépares les matrices."""
    dfs_requis = []
    dfs_autres = {}
    labels = []

    for nom in matrices_cibles:
        logging.info(" Chargement de la matrice %s", nom)
        df = loader.get_matrix(nom)
        df = nettoyer_colonnes(df)

        remap = loader.get_column_mapping(nom)
        if remap:
            logging.info(" Remapping détecté : %s", remap)
            df = df.rename(columns=remap)

        if not all(col in df.columns for col in KEY_COLS):
            logging.warning(" Colonnes clés manquantes dans %s", nom)
            continue

        df_requis, df_non_requis = separer_requis(df)
        logging.info(
            "%s contient %s requis et %s non-requis.",
            nom,
            len(df_requis),
            len(df_non_requis),
        )

        dfs_requis.append(df_requis)
        dfs_autres[nom] = df_non_requis
        labels.append(nom)

    return dfs_requis, dfs_autres, labels


def log_summary(summary):
    """Logue les infos."""
    logging.info("\n--- Résumé ---")
    for k, v in summary.items():
        logging.info("%s: %s", k, v)


def exporter_resultats(output_dir, exclusifs, commun, divergents, labels):
    """Exporte les résultats en Excel."""
    commun.to_excel(f"{output_dir}/entries_common_to_all.xlsx", index=False)
    for nom, df in exclusifs.items():
        df.to_excel(f"{output_dir}/entries_unique_to_{nom}.xlsx", index=False)

    if divergents is not None and not divergents.empty:
        nom_concat = "-".join(labels)
        divergents.to_excel(f"{output_dir}/comparison_{nom_concat}.xlsx", index=False)


def analyser_si_divergences(divergents, output_dir, labels):
    """Analyses les divergences."""
    if divergents is not None and not divergents.empty:
        requis_impactes = set(divergents["Reference"])
        logging.info(" Nombre de requis impactés : %s", len(requis_impactes))
        res = analyser_divergences_documentaires(divergents, source_cols=labels)
        res_filen = f"{output_dir}/res_ana_div_{'-'.join(labels)}.xlsx"
        res.to_excel(res_filen, index=False)
        logging.info(" Analyse documentaire enregistrée dans %s", res_filen)
        return res

    logging.info(" Aucun champ divergent détecté.")
    return None
